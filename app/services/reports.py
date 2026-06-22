"""
Rapor motoru: PDF ve Excel çıktısı üretir.
"""
import io
import os
from datetime import datetime
from collections import defaultdict
from typing import Optional

from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.models import Project, Output, User, OutputType, OutputStatus, UserRole


# ── Font yönetimi ────────────────────────────────────────────────────────────

def _find_unicode_font() -> Optional[str]:
    candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _register_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    path = _find_unicode_font()
    if not path:
        return "Helvetica", "Helvetica-Bold"

    base = os.path.splitext(path)[0]
    bold_candidates = [
        base + "b.ttf", base + "bd.ttf", base + "-Bold.ttf", base + "Bold.ttf",
    ]
    bold_path = next((p for p in bold_candidates if os.path.exists(p)), None)

    try:
        pdfmetrics.registerFont(TTFont("TRFont", path))
        if bold_path:
            pdfmetrics.registerFont(TTFont("TRFontBold", bold_path))
            return "TRFont", "TRFontBold"
        return "TRFont", "TRFont"
    except Exception:
        return "Helvetica", "Helvetica-Bold"


# ── Veri toplama ─────────────────────────────────────────────────────────────

def _get_report_data(db: Session, year: Optional[int] = None) -> dict:
    query = db.query(Output).join(Project)
    if year:
        query = query.filter(Project.application_year == year)
    outputs = query.filter(Output.status == OutputStatus.onaylandi).all()

    by_type: dict     = defaultdict(int)
    by_faculty: dict  = defaultdict(int)
    by_dept: dict     = defaultdict(int)
    by_person: dict   = defaultdict(lambda: {"count": 0, "faculty": "", "department": "", "types": defaultdict(int)})

    for o in outputs:
        by_type[o.output_type.value] += 1
        pi = o.project.principal_investigator
        fac = (pi.faculty or "Belirtilmemiş") if pi else "Belirtilmemiş"
        dep = (pi.department or o.project.department or "Belirtilmemiş") if pi else (o.project.department or "Belirtilmemiş")
        by_faculty[fac] += 1
        by_dept[dep] += 1

        submitter = o.submitted_by
        key = submitter.full_name if submitter else "Bilinmeyen"
        by_person[key]["count"] += 1
        by_person[key]["faculty"] = (submitter.faculty or fac) if submitter else fac
        by_person[key]["department"] = (submitter.department or dep) if submitter else dep
        by_person[key]["types"][o.output_type.value] += 1

    proj_q = db.query(Project).filter(Project.is_active == True)
    if year:
        proj_q = proj_q.filter(Project.application_year == year)
    all_projects = proj_q.order_by(Project.project_code).all()

    # Kişi-proje tablosu için tüm faculty kullanıcıları
    researchers = db.query(User).filter(User.role == UserRole.faculty, User.is_active == True).all()
    researcher_rows = []
    for u in researchers:
        led = [p for p in u.led_projects if not year or p.application_year == year]
        all_proj = list({p.id: p for p in led + list(u.research_projects)}.values())
        if year:
            all_proj = [p for p in all_proj if p.application_year == year]
        approved_outputs = [o for p in all_proj for o in p.outputs if o.status == OutputStatus.onaylandi]
        type_counts: dict = defaultdict(int)
        for o in approved_outputs:
            type_counts[o.output_type.value] += 1
        researcher_rows.append({
            "name": u.full_name,
            "faculty": u.faculty or "—",
            "department": u.department or "—",
            "project_count": len(all_proj),
            "output_count": len(approved_outputs),
            "types": dict(type_counts),
            "projects": all_proj,
        })
    researcher_rows.sort(key=lambda x: -x["output_count"])

    approved_project_ids = set(o.project_id for o in outputs)
    no_output_projects = []
    no_output_declared_projects = []
    for p in all_projects:
        if p.id in approved_project_ids:
            continue
        pi = p.principal_investigator
        entry = {
            "code":       p.project_code,
            "title":      p.title,
            "type":       p.project_type or "—",
            "pi":         pi.full_name if pi else "—",
            "faculty":    (pi.faculty or "—") if pi else "—",
            "department": p.department or (pi.department if pi else "") or "—",
            "budget":     f"{p.budget:,.0f}" if p.budget else "—",
            "pending":    sum(1 for o in p.outputs if o.status == OutputStatus.gonderildi),
            "declared":   p.no_output_declared,
        }
        if p.no_output_declared:
            no_output_declared_projects.append(entry)
        else:
            no_output_projects.append(entry)

    return {
        "outputs": outputs,
        "by_type": dict(by_type),
        "by_faculty": dict(by_faculty),
        "by_department": dict(by_dept),
        "by_person": dict(by_person),
        "total_outputs": len(outputs),
        "projects_with_output": len(approved_project_ids),
        "total_projects": len(all_projects),
        "all_projects": all_projects,
        "researcher_rows": researcher_rows,
        "no_output_projects": no_output_projects,
        "no_output_declared_projects": no_output_declared_projects,
        "year": year,
    }


# ── PDF Raporu ────────────────────────────────────────────────────────────────

def generate_pdf_report(db: Session, year: Optional[int] = None) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether,
    )

    fn, fnb = _register_font()
    data = _get_report_data(db, year)
    buf = io.BytesIO()
    _ns = str(id(buf))

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm, bottomMargin=2.5*cm,
        title=f"BAP Çıktı Raporu {data['year'] or 'Tüm Yıllar'}",
        author="BAP Koordinasyon Birimi",
    )

    # Renkler
    C_NAVY   = colors.HexColor("#1A365D")
    C_BLUE   = colors.HexColor("#2B6CB0")
    C_TEAL   = colors.HexColor("#2C7A7B")
    C_PURPLE = colors.HexColor("#553C9A")
    C_ORANGE = colors.HexColor("#C05621")
    C_GREEN  = colors.HexColor("#276749")
    C_GRAY   = colors.HexColor("#4A5568")
    C_LGRAY  = colors.HexColor("#F7FAFC")
    C_LINE   = colors.HexColor("#CBD5E0")
    C_GOLD   = colors.HexColor("#B7791F")

    def S(name, **kw):
        kw.setdefault("fontName", fn)
        kw.setdefault("fontSize", 10)
        kw.setdefault("leading", 15)
        return ParagraphStyle(f"{name}_{_ns}", **kw)

    s_title    = S("t",  fontName=fnb, fontSize=22, textColor=C_NAVY,   spaceAfter=4)
    s_subtitle = S("st", fontSize=11,  textColor=C_GRAY,  spaceAfter=20)
    s_h2       = S("h2", fontName=fnb, fontSize=14, textColor=C_NAVY,   spaceBefore=20, spaceAfter=8)
    s_h3       = S("h3", fontName=fnb, fontSize=11, textColor=C_BLUE,   spaceBefore=12, spaceAfter=4)
    s_body     = S("b",  fontSize=10,  leading=15)
    s_small    = S("s",  fontSize=8.5, textColor=C_GRAY,  leading=13)
    s_cell     = S("c",  fontSize=8,   leading=12)
    s_cell_b   = S("cb", fontName=fnb, fontSize=8, leading=12)
    s_footer   = S("f",  fontSize=8,   textColor=C_GRAY)
    s_rank     = S("rk", fontName=fnb, fontSize=11, textColor=C_GOLD, leading=14)

    def tbl(rows, widths, hdr_color=C_BLUE, font_size=9):
        t = Table(rows, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  hdr_color),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  fnb),
            ("FONTNAME",       (0, 1), (-1, -1), fn),
            ("FONTSIZE",       (0, 0), (-1, -1), font_size),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_LGRAY]),
            ("BOX",            (0, 0), (-1, -1), 0.5, C_LINE),
            ("INNERGRID",      (0, 0), (-1, -1), 0.3, C_LINE),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
            ("LEFTPADDING",    (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
            ("VALIGN",         (0, 0), (-1, -1), "TOP"),
        ]))
        return t

    def rank_label(i):
        return {1: "1.", 2: "2.", 3: "3."}.get(i, f"{i}.")

    def bar_str(pct, width=15):
        filled = int(pct / 100 * width)
        return "|" * filled + "." * (width - filled)

    story = []

    # ══ KAPAK ══
    story.append(Paragraph("BAP Proje Ciktilari", s_title))
    story.append(Paragraph(
        f"Ayrintili Analiz Raporu  -  "
        f"{'<b>' + str(data['year']) + ' Basvuru Yili</b>' if data['year'] else '<b>Tum Yillar</b>'}  -  "
        f"Olusturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        s_subtitle,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=C_BLUE))
    story.append(Spacer(1, 16))

    # ══ GENEL ÖZET ══
    story.append(Paragraph("1. Genel Ozet", s_h2))
    total_out  = data["total_outputs"]
    total_proj = data["total_projects"]
    with_out   = data["projects_with_output"]
    without    = total_proj - with_out
    rate       = f"{with_out / total_proj * 100:.1f}%" if total_proj else "-"
    avg_per    = f"{total_out / with_out:.1f}" if with_out else "-"

    summary_rows = [
        ["Gosterge", "Deger", "Aciklama"],
        ["Toplam Onayli Cikti",       str(total_out),  "Sistemde onaylanmis tum ciktilar"],
        ["Ciktisi Olan Proje",        str(with_out),   f"Toplam {total_proj} projeden {rate} oraninda"],
        ["Ciktisi Olmayan Proje",     str(without),    "Henuz cikti girilmemis projeler"],
        ["Proje Basina Ort. Cikti",   avg_per,         "Ciktisi olan projeler icin ortalama"],
        ["Aktif Arastirmaci",         str(len(data["researcher_rows"])), "Sistemdeki toplam akademisyen"],
    ]
    story.append(tbl(summary_rows, [8*cm, 3*cm, 6*cm], hdr_color=C_NAVY))
    story.append(Spacer(1, 20))

    # ══ FAKÜLTELERİ SIRAL ══
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LINE))
    story.append(Paragraph("2. Fakulte Bazli Cikti Siralamasi", s_h2))
    if data["by_faculty"]:
        fac_sorted = sorted(data["by_faculty"].items(), key=lambda x: -x[1])
        fac_rows = [["Sira", "Fakulte", "Cikti Sayisi", "Oran (%)", "Gosterge"]]
        for i, (fac, cnt) in enumerate(fac_sorted, 1):
            pct = cnt / total_out * 100 if total_out else 0
            fac_rows.append([rank_label(i), fac, str(cnt), f"{pct:.1f}%", bar_str(pct)])
        story.append(tbl(fac_rows, [1.2*cm, 7.5*cm, 2.5*cm, 2*cm, 3.8*cm], hdr_color=C_TEAL))
    story.append(Spacer(1, 20))

    # ══ BÖLÜMLERİ SIRAL ══
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LINE))
    story.append(Paragraph("3. Bolum Bazli Cikti Siralamasi", s_h2))
    if data["by_department"]:
        dep_sorted = sorted(data["by_department"].items(), key=lambda x: -x[1])
        dep_rows = [["Sira", "Bolum", "Cikti Sayisi", "Oran (%)", "Gosterge"]]
        for i, (dep, cnt) in enumerate(dep_sorted, 1):
            pct = cnt / total_out * 100 if total_out else 0
            dep_rows.append([rank_label(i), dep, str(cnt), f"{pct:.1f}%", bar_str(pct)])
        story.append(tbl(dep_rows, [1.2*cm, 7.5*cm, 2.5*cm, 2*cm, 3.8*cm], hdr_color=C_TEAL))
    story.append(Spacer(1, 20))

    # ══ PROJE TÜRÜNE GÖRE ÇIKTI ══
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LINE))
    story.append(Paragraph("4. Cikti Turu Dagilimi", s_h2))
    if data["by_type"]:
        type_sorted = sorted(data["by_type"].items(), key=lambda x: -x[1])
        type_rows = [["Sira", "Cikti Turu", "Adet", "Oran (%)", "Gosterge"]]
        for i, (otype, cnt) in enumerate(type_sorted, 1):
            pct = cnt / total_out * 100 if total_out else 0
            type_rows.append([rank_label(i), otype, str(cnt), f"{pct:.1f}%", bar_str(pct)])
        story.append(tbl(type_rows, [1.2*cm, 5.5*cm, 2*cm, 2.3*cm, 6*cm], hdr_color=C_PURPLE))
    story.append(Spacer(1, 20))

    # ══ EN ÇOK ÇIKTI ÜRETEN KİŞİLER ══
    story.append(PageBreak())
    story.append(Paragraph("5. En Cok Cikti Ureten Arastirmacilar", s_h2))

    if data["researcher_rows"]:
        top_rows = [["Sira", "Ad Soyad", "Fakulte / Bolum", "Proje", "Toplam", "Makale", "Bildiri", "Patent", "Diger"]]
        for i, r in enumerate(data["researcher_rows"][:30], 1):
            if r["output_count"] == 0:
                continue
            fac_dep = r["faculty"]
            if r["department"] and r["department"] != "-":
                fac_dep += f" / {r['department']}"
            other = r["output_count"] - r["types"].get("Makale/Yayin", 0) - r["types"].get("Bildiri", 0) - r["types"].get("Patent", 0)
            other = max(other, 0)
            top_rows.append([
                rank_label(i),
                Paragraph(r["name"], s_cell_b),
                Paragraph(fac_dep, s_cell),
                str(r["project_count"]),
                str(r["output_count"]),
                str(r["types"].get("Makale/Yayin", 0) or r["types"].get("Makale/Yayın", 0)),
                str(r["types"].get("Bildiri", 0)),
                str(r["types"].get("Patent", 0)),
                str(other),
            ])
        story.append(tbl(top_rows, [1.2*cm, 4*cm, 4.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.4*cm], hdr_color=C_ORANGE))

    story.append(Spacer(1, 20))

    # ══ KİŞİ-PROJE KARŞILAŞTIRMA TABLOSU ══
    story.append(PageBreak())
    story.append(Paragraph("6. Arastirmaci - Proje - Cikti Detay Tablosu", s_h2))
    story.append(Paragraph(
        "Her arastirmacinin yuruttuugu projeler ve bu projelere ait onayli ciktilar asagida listelenmektedir.",
        s_small,
    ))
    story.append(Spacer(1, 8))

    for r in data["researcher_rows"]:
        if r["output_count"] == 0:
            continue

        block = []
        block.append(Paragraph(
            f"<b>{r['name']}</b>  -  {r['faculty']}  /  {r['department']}  "
            f"-  <b>{r['output_count']}</b> cikti  -  {r['project_count']} proje",
            s_h3,
        ))

        for proj in r["projects"]:
            approved = [o for o in proj.outputs if o.status == OutputStatus.onaylandi]
            if not approved:
                continue

            block.append(Paragraph(
                f"  <b>{proj.project_code}</b>  -  {proj.title}"
                + (f"  [{proj.project_type}]" if proj.project_type else ""),
                s_small,
            ))
            out_rows = [["Tur", "Cikti Basligi", "Yazar(lar)", "Yayinci / Dergi", "Tarih", "DOI/ISBN"]]
            for o in approved:
                out_rows.append([
                    Paragraph(o.output_type.value, s_cell),
                    Paragraph(o.title or "-", s_cell),
                    Paragraph(o.authors or "-", s_cell),
                    Paragraph(o.publisher_venue or "-", s_cell),
                    Paragraph(o.publication_date or "-", s_cell),
                    Paragraph(o.identifier or "-", s_cell),
                ])
            t = Table(out_rows, colWidths=[2.3*cm, 4.5*cm, 3*cm, 3*cm, 1.8*cm, 2.4*cm], repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#EBF8FF")),
                ("TEXTCOLOR",      (0, 0), (-1, 0),  C_NAVY),
                ("FONTNAME",       (0, 0), (-1, 0),  fnb),
                ("FONTNAME",       (0, 1), (-1, -1), fn),
                ("FONTSIZE",       (0, 0), (-1, -1), 7.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_LGRAY]),
                ("BOX",            (0, 0), (-1, -1), 0.4, C_LINE),
                ("INNERGRID",      (0, 0), (-1, -1), 0.3, C_LINE),
                ("TOPPADDING",     (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
                ("LEFTPADDING",    (0, 0), (-1, -1), 4),
                ("VALIGN",         (0, 0), (-1, -1), "TOP"),
            ]))
            block.append(t)
            block.append(Spacer(1, 4))

        block.append(Spacer(1, 8))
        story.append(KeepTogether(block))

    # ══ ÇIKTISI OLMAYAN PROJELER ══
    story.append(PageBreak())
    story.append(Paragraph("7. Ciktisi Olmayan Projeler", s_h2))

    C_RED = colors.HexColor("#C53030")
    C_RED_L = colors.HexColor("#FFF5F5")

    no_out_cols = ["Proje Kodu", "Proje Basligi", "Tur", "Yurutucu", "Fakulte / Bolum", "Butce (TL)", "Durum"]

    if data["no_output_projects"]:
        story.append(Paragraph(
            f"Asagidaki {len(data['no_output_projects'])} projenin henuz onayli ciktisi yoktur.",
            s_small,
        ))
        story.append(Spacer(1, 6))
        rows_no = [no_out_cols]
        for e in data["no_output_projects"]:
            durum = f"Beklemede ({e['pending']})" if e["pending"] else "Hic Eklenmemis"
            rows_no.append([
                Paragraph(e["code"], s_cell),
                Paragraph(e["title"], s_cell),
                Paragraph(e["type"], s_cell),
                Paragraph(e["pi"], s_cell),
                Paragraph(f"{e['faculty']} / {e['department']}", s_cell),
                Paragraph(e["budget"], s_cell),
                Paragraph(durum, s_cell),
            ])
        t = Table(rows_no, colWidths=[2.2*cm, 5*cm, 1.8*cm, 3*cm, 3.5*cm, 2*cm, 2.5*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  C_RED),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  fnb),
            ("FONTNAME",       (0, 1), (-1, -1), fn),
            ("FONTSIZE",       (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_RED_L]),
            ("BOX",            (0, 0), (-1, -1), 0.4, C_LINE),
            ("INNERGRID",      (0, 0), (-1, -1), 0.3, C_LINE),
            ("TOPPADDING",     (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
            ("LEFTPADDING",    (0, 0), (-1, -1), 4),
            ("VALIGN",         (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Tum projelerin en az bir onayli ciktisi bulunmaktadir.", s_small))

    story.append(Spacer(1, 16))

    if data["no_output_declared_projects"]:
        story.append(Paragraph(
            f"Asagidaki {len(data['no_output_declared_projects'])} proje icin 'Cikti Bulunmamaktadir' bildirimi yapilmistir.",
            s_small,
        ))
        story.append(Spacer(1, 6))
        C_PURPLE2 = colors.HexColor("#553C9A")
        C_PURPLE_L = colors.HexColor("#FAF5FF")
        rows_decl = [no_out_cols]
        for e in data["no_output_declared_projects"]:
            rows_decl.append([
                Paragraph(e["code"], s_cell),
                Paragraph(e["title"], s_cell),
                Paragraph(e["type"], s_cell),
                Paragraph(e["pi"], s_cell),
                Paragraph(f"{e['faculty']} / {e['department']}", s_cell),
                Paragraph(e["budget"], s_cell),
                Paragraph("Cikti Yok Bildirimi", s_cell),
            ])
        t2 = Table(rows_decl, colWidths=[2.2*cm, 5*cm, 1.8*cm, 3*cm, 3.5*cm, 2*cm, 2.5*cm], repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  C_PURPLE2),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  fnb),
            ("FONTNAME",       (0, 1), (-1, -1), fn),
            ("FONTSIZE",       (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_PURPLE_L]),
            ("BOX",            (0, 0), (-1, -1), 0.4, C_LINE),
            ("INNERGRID",      (0, 0), (-1, -1), 0.3, C_LINE),
            ("TOPPADDING",     (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
            ("LEFTPADDING",    (0, 0), (-1, -1), 4),
            ("VALIGN",         (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t2)

    # ══ PROJE LİSTESİ ══
    story.append(PageBreak())
    story.append(Paragraph("8. Proje Bazli Cikti Ozeti", s_h2))
    proj_rows = [["Proje Kodu", "Baslik", "Tur", "Yurutucu", "Bolum", "Butce (TL)", "Basl.", "Bitis", "Cikti"]]
    for proj in data["all_projects"]:
        approved_cnt = sum(1 for o in proj.outputs if o.status == OutputStatus.onaylandi)
        pi = proj.principal_investigator
        proj_rows.append([
            Paragraph(proj.project_code, s_cell),
            Paragraph(proj.title, s_cell),
            Paragraph(proj.project_type or "-", s_cell),
            Paragraph(pi.full_name if pi else "-", s_cell),
            Paragraph(proj.department or (pi.department if pi else "") or "-", s_cell),
            Paragraph(f"{proj.budget:,.0f}" if proj.budget else "-", s_cell),
            Paragraph(proj.start_date.strftime("%d.%m.%Y") if proj.start_date else "-", s_cell),
            Paragraph(proj.end_date.strftime("%d.%m.%Y") if proj.end_date else "-", s_cell),
            str(approved_cnt),
        ])
    story.append(tbl(
        proj_rows,
        [2.5*cm, 4.5*cm, 2*cm, 3*cm, 2.5*cm, 2*cm, 1.8*cm, 1.8*cm, 1.2*cm],
        hdr_color=C_GREEN, font_size=8,
    ))

    # ══ ALT BİLGİ ══
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LINE))
    story.append(Paragraph(
        f"Bu rapor {datetime.now().strftime('%d.%m.%Y %H:%M')} tarihinde BAP Cikti Yonetim Sistemi tarafindan otomatik olusturulmustur.  "
        f"Kapsam: {'Basvuru yili ' + str(data['year']) if data['year'] else 'Tum yillar'}.",
        s_footer,
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Excel Raporu ──────────────────────────────────────────────────────────────

def generate_excel_report(db: Session, year: Optional[int] = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

    data = _get_report_data(db, year)
    wb = openpyxl.Workbook()

    # Stil sabitleri
    NAVY   = "1A365D"
    BLUE   = "2B6CB0"
    TEAL   = "2C7A7B"
    PURPLE = "553C9A"
    ORANGE = "C05621"
    GREEN  = "276749"
    LGRAY  = "F7FAFC"
    LINE   = "CBD5E0"

    def hfill(color):
        return PatternFill("solid", fgColor=color)

    def hfont(sz=10):
        return Font(color="FFFFFF", bold=True, size=sz)

    def bold(sz=10, color="000000"):
        return Font(bold=True, size=sz, color=color)

    TITLE_FONT = Font(bold=True, size=15, color=NAVY)
    LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    TOP_C  = Alignment(horizontal="center", vertical="top",    wrap_text=True)
    thin = Side(style="thin", color=LINE)
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hrow(ws, row, cols, color=BLUE):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font  = hfont()
            cell.fill  = hfill(color)
            cell.alignment = CENTER
            cell.border = BORDER

    def dcell(ws, row, col, val, align=LEFT, fmt=None, num_color=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = align
        cell.border    = BORDER
        if fmt:
            cell.number_format = fmt
        if num_color and isinstance(val, (int, float)) and val > 0:
            cell.font = Font(bold=True, color=num_color)
        return cell

    def title_row(ws, row, text, span_end, color=NAVY):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = TITLE_FONT
        c.fill      = hfill("EBF8FF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER
        ws.row_dimensions[row].height = 28

    def section_label(ws, row, text, col=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=11, color=NAVY)
        ws.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════
    # SAYFA 1 — ÖZET & SIRALAMALAR
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Ozet ve Siralamalar"
    year_label = f"{data['year']} Başvuru Yılı" if data["year"] else "Tüm Yıllar"

    title_row(ws1, 1, f"BAP Proje Çıktıları — {year_label} — Özet Rapor", 8)
    ws1.cell(row=2, column=1, value=f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(italic=True, color="718096")

    # Genel istatistikler
    section_label(ws1, 4, "GENEL İSTATİSTİKLER")
    hrow(ws1, 5, ["Gösterge", "Değer", "Not"], color=NAVY)
    total_out  = data["total_outputs"]
    total_proj = data["total_projects"]
    with_out   = data["projects_with_output"]
    stat_rows = [
        ("Toplam Onaylı Çıktı",       total_out,              "—"),
        ("Çıktısı Olan Proje",        with_out,               f"{with_out/total_proj*100:.1f}% (/{total_proj} proje)" if total_proj else "—"),
        ("Çıktısı Olmayan Proje",     total_proj - with_out,  "Henüz çıktı eklenmemiş"),
        ("Proje Başına Ort. Çıktı",   round(total_out/with_out,2) if with_out else 0, "Çıktılı projeler için"),
        ("Toplam Aktif Araştırmacı",  len(data["researcher_rows"]), "Sistemdeki fakulti kullanıcıları"),
    ]
    for i, (label, val, note) in enumerate(stat_rows, 6):
        dcell(ws1, i, 1, label)
        dcell(ws1, i, 2, val, CENTER, num_color=BLUE)
        dcell(ws1, i, 3, note)

    # Türe göre sıralama
    r = 13
    section_label(ws1, r, "ÇIKTI TÜRÜ SIRALAMASI (Yüksekten Düşüğe)")
    r += 1
    hrow(ws1, r, ["Sıra", "Çıktı Türü", "Adet", "Oran (%)", "Kümülatif (%)"], color=PURPLE)
    r += 1
    type_sorted = sorted(data["by_type"].items(), key=lambda x: -x[1])
    cumul = 0
    for i, (otype, cnt) in enumerate(type_sorted, 1):
        pct = cnt / total_out * 100 if total_out else 0
        cumul += pct
        dcell(ws1, r, 1, i, CENTER)
        dcell(ws1, r, 2, otype)
        dcell(ws1, r, 3, cnt, CENTER, num_color=PURPLE)
        dcell(ws1, r, 4, round(pct, 1), CENTER)
        dcell(ws1, r, 5, round(cumul, 1), CENTER)
        r += 1

    # Fakülteye göre sıralama
    r += 1
    section_label(ws1, r, "FAKÜLTELERİN ÇIKTI SIRALAMASI")
    r += 1
    hrow(ws1, r, ["Sıra", "Fakülte", "Çıktı Sayısı", "Oran (%)", "Kümülatif (%)", "Bar"], color=TEAL)
    r += 1
    fac_sorted = sorted(data["by_faculty"].items(), key=lambda x: -x[1])
    cumul = 0
    max_fac = fac_sorted[0][1] if fac_sorted else 1
    for i, (fac, cnt) in enumerate(fac_sorted, 1):
        pct = cnt / total_out * 100 if total_out else 0
        cumul += pct
        bar_len = int(cnt / max_fac * 20)
        dcell(ws1, r, 1, i, CENTER)
        dcell(ws1, r, 2, fac)
        dcell(ws1, r, 3, cnt, CENTER, num_color=TEAL)
        dcell(ws1, r, 4, round(pct, 1), CENTER)
        dcell(ws1, r, 5, round(cumul, 1), CENTER)
        dcell(ws1, r, 6, "█" * bar_len)
        r += 1

    # Bölüme göre sıralama
    r += 1
    section_label(ws1, r, "BÖLÜMLERİN ÇIKTI SIRALAMASI")
    r += 1
    hrow(ws1, r, ["Sıra", "Bölüm", "Çıktı Sayısı", "Oran (%)", "Kümülatif (%)", "Bar"], color=TEAL)
    r += 1
    dep_sorted = sorted(data["by_department"].items(), key=lambda x: -x[1])
    cumul = 0
    max_dep = dep_sorted[0][1] if dep_sorted else 1
    for i, (dep, cnt) in enumerate(dep_sorted, 1):
        pct = cnt / total_out * 100 if total_out else 0
        cumul += pct
        bar_len = int(cnt / max_dep * 20)
        dcell(ws1, r, 1, i, CENTER)
        dcell(ws1, r, 2, dep)
        dcell(ws1, r, 3, cnt, CENTER, num_color=TEAL)
        dcell(ws1, r, 4, round(pct, 1), CENTER)
        dcell(ws1, r, 5, round(cumul, 1), CENTER)
        dcell(ws1, r, 6, "█" * bar_len)
        r += 1

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 38
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 22
    ws1.freeze_panes = "A6"

    # ══════════════════════════════════════════════
    # SAYFA 2 — EN ÇOK ÇIKTI ÜRETEN ARAŞTIRMACLAR
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Arastirmaci Siralamasi")
    title_row(ws2, 1, f"En Çok Çıktı Üreten Araştırmacılar — {year_label}", 10)
    ws2.cell(row=2, column=1, value="Sadece en az 1 onaylı çıktıya sahip araştırmacılar listelenmiştir.").font = Font(italic=True, color="718096")

    hrow(ws2, 3, [
        "Sıra", "Ad Soyad", "Fakülte", "Bölüm",
        "Proje Sayısı", "Toplam Çıktı", "Makale/Yayın",
        "Bildiri", "Patent", "Diğer",
    ], color=ORANGE)

    r = 4
    for i, row in enumerate(data["researcher_rows"], 1):
        if row["output_count"] == 0:
            continue
        other = row["output_count"] - row["types"].get("Makale/Yayın", 0) - row["types"].get("Bildiri", 0) - row["types"].get("Patent", 0)
        fill_color = "FFF3E0" if i <= 3 else ("FFFFFF" if i % 2 == 0 else "F7FAFC")
        for col in range(1, 11):
            c = ws2.cell(row=r, column=col)
            c.fill   = PatternFill("solid", fgColor=fill_color)
            c.border = BORDER

        dcell(ws2, r, 1, i, CENTER)
        dcell(ws2, r, 2, row["name"])
        dcell(ws2, r, 3, row["faculty"])
        dcell(ws2, r, 4, row["department"])
        dcell(ws2, r, 5, row["project_count"], CENTER)
        cnt_cell = dcell(ws2, r, 6, row["output_count"], CENTER)
        cnt_cell.font = Font(bold=True, size=11, color=ORANGE)
        dcell(ws2, r, 7, row["types"].get("Makale/Yayın", 0), CENTER)
        dcell(ws2, r, 8, row["types"].get("Bildiri", 0), CENTER)
        dcell(ws2, r, 9, row["types"].get("Patent", 0), CENTER)
        dcell(ws2, r, 10, other, CENTER)
        r += 1

    col_ws2 = [6, 28, 28, 25, 12, 12, 14, 10, 10, 10]
    for i, w in enumerate(col_ws2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"
    ws2.row_dimensions[3].height = 30

    # ══════════════════════════════════════════════
    # SAYFA 3 — KİŞİ-PROJE KARŞILAŞTIRMA
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Kisi-Proje Detayi")
    title_row(ws3, 1, f"Araştırmacı — Proje — Çıktı Karşılaştırma Tablosu — {year_label}", 12)
    hrow(ws3, 2, [
        "Araştırmacı", "Fakülte", "Bölüm",
        "Proje Kodu", "Proje Başlığı", "Proje Türü",
        "Bütçe (₺)", "Başlangıç", "Bitiş",
        "Proje Çıktı", "Makale", "Bildiri",
    ], color=NAVY)

    r = 3
    for row in data["researcher_rows"]:
        name_written = False
        for proj in row["projects"]:
            approved = [o for o in proj.outputs if o.status == OutputStatus.onaylandi]
            type_counts: dict = defaultdict(int)
            for o in approved:
                type_counts[o.output_type.value] += 1

            fill = "F0FFF4" if not name_written else ("FFFFFF" if r % 2 == 0 else "F7FAFC")
            for col in range(1, 13):
                ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)

            dcell(ws3, r, 1, row["name"] if not name_written else "")
            dcell(ws3, r, 2, row["faculty"] if not name_written else "")
            dcell(ws3, r, 3, row["department"] if not name_written else "")
            dcell(ws3, r, 4, proj.project_code, CENTER)
            dcell(ws3, r, 5, proj.title)
            dcell(ws3, r, 6, proj.project_type or "—")
            budget_c = dcell(ws3, r, 7, proj.budget or 0, CENTER, fmt='#,##0.00')
            dcell(ws3, r, 8, proj.start_date.strftime("%d.%m.%Y") if proj.start_date else "—", CENTER)
            dcell(ws3, r, 9,  proj.end_date.strftime("%d.%m.%Y") if proj.end_date else "—", CENTER)
            out_c = dcell(ws3, r, 10, len(approved), CENTER)
            if len(approved) > 0:
                out_c.font = Font(bold=True, color=GREEN)
            dcell(ws3, r, 11, type_counts.get("Makale/Yayın", 0), CENTER)
            dcell(ws3, r, 12, type_counts.get("Bildiri", 0), CENTER)

            name_written = True
            r += 1

        if not row["projects"]:
            dcell(ws3, r, 1, row["name"])
            dcell(ws3, r, 2, row["faculty"])
            dcell(ws3, r, 3, row["department"])
            for col in range(4, 13):
                dcell(ws3, r, col, "—", CENTER)
            r += 1

    col_ws3 = [24, 26, 24, 14, 36, 16, 14, 13, 13, 12, 10, 10]
    for i, w in enumerate(col_ws3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A3"
    ws3.row_dimensions[2].height = 30

    # ══════════════════════════════════════════════
    # SAYFA 4 — TÜM ÇIKTILAR (DETAY)
    # ══════════════════════════════════════════════
    ws4 = wb.create_sheet("Tum Ciktilar")
    title_row(ws4, 1, f"Onaylı Çıktıların Tam Listesi — {year_label}", 15)
    cols4 = [
        "Proje Kodu", "Proje Başlığı", "Proje Türü", "Bölüm (Proje)",
        "Yürütücü", "Fakülte", "Bölüm (Yürütücü)",
        "Çıktı Türü", "Çıktı Başlığı",
        "Yazar(lar)", "Yayın Tarihi",
        "DOI / ISBN / Patent No", "Yayıncı / Dergi / Konferans",
        "BAP Atıf Notu", "Eklenme Tarihi",
    ]
    hrow(ws4, 2, cols4, color=GREEN)
    ws4.freeze_panes = "A3"

    for ri, o in enumerate(data["outputs"], start=3):
        pi = o.project.principal_investigator
        fill = "FFFFFF" if ri % 2 == 0 else "F7FAFC"
        for col in range(1, 16):
            ws4.cell(row=ri, column=col).fill = PatternFill("solid", fgColor=fill)
        dcell(ws4, ri,  1, o.project.project_code, CENTER)
        dcell(ws4, ri,  2, o.project.title)
        dcell(ws4, ri,  3, o.project.project_type or "—")
        dcell(ws4, ri,  4, o.project.department or "—")
        dcell(ws4, ri,  5, pi.full_name if pi else "—")
        dcell(ws4, ri,  6, (pi.faculty or "—") if pi else "—")
        dcell(ws4, ri,  7, (pi.department or "—") if pi else "—")
        dcell(ws4, ri,  8, o.output_type.value, CENTER)
        dcell(ws4, ri,  9, o.title)
        dcell(ws4, ri, 10, o.authors or "—")
        dcell(ws4, ri, 11, o.publication_date or "—", CENTER)
        dcell(ws4, ri, 12, o.identifier or "—")
        dcell(ws4, ri, 13, o.publisher_venue or "—")
        dcell(ws4, ri, 14, o.acknowledgement_note or "—")
        dcell(ws4, ri, 15, o.created_at.strftime("%d.%m.%Y"), CENTER)

    col_ws4 = [14, 32, 16, 20, 22, 22, 20, 18, 36, 28, 14, 22, 28, 30, 14]
    for i, w in enumerate(col_ws4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.row_dimensions[2].height = 32

    # ══════════════════════════════════════════════
    # SAYFA 5 — PROJE BAZLI ÖZET
    # ══════════════════════════════════════════════
    ws5 = wb.create_sheet("Proje Bazli")
    title_row(ws5, 1, f"Proje Bazlı Çıktı Özeti — {year_label}", 14)
    cols5 = [
        "Proje Kodu", "Proje Başlığı", "Proje Türü", "Bölüm",
        "Yürütücü", "Fakülte", "Bütçe (₺)",
        "Başlangıç", "Bitiş",
        "Toplam Çıktı", "Makale", "Bildiri", "Patent", "Diğer",
    ]
    hrow(ws5, 2, cols5, color=NAVY)
    ws5.freeze_panes = "A3"

    for ri, proj in enumerate(data["all_projects"], start=3):
        approved = [o for o in proj.outputs if o.status == OutputStatus.onaylandi]
        tc: dict = defaultdict(int)
        for o in approved:
            tc[o.output_type.value] += 1
        pi  = proj.principal_investigator
        oth = sum(v for k, v in tc.items() if k not in ("Makale/Yayın", "Bildiri", "Patent"))
        fill = "FFFFFF" if ri % 2 == 0 else "F7FAFC"
        for col in range(1, 15):
            ws5.cell(row=ri, column=col).fill = PatternFill("solid", fgColor=fill)
        dcell(ws5, ri,  1, proj.project_code, CENTER)
        dcell(ws5, ri,  2, proj.title)
        dcell(ws5, ri,  3, proj.project_type or "—")
        dcell(ws5, ri,  4, proj.department or (pi.department if pi else "") or "—")
        dcell(ws5, ri,  5, pi.full_name if pi else "—")
        dcell(ws5, ri,  6, (pi.faculty or "—") if pi else "—")
        bc = dcell(ws5, ri, 7, proj.budget or 0, CENTER, fmt='#,##0.00')
        dcell(ws5, ri,  8, proj.start_date.strftime("%d.%m.%Y") if proj.start_date else "—", CENTER)
        dcell(ws5, ri,  9, proj.end_date.strftime("%d.%m.%Y") if proj.end_date else "—", CENTER)
        oc = dcell(ws5, ri, 10, len(approved), CENTER)
        if len(approved) > 0:
            oc.font = Font(bold=True, color=GREEN)
        dcell(ws5, ri, 11, tc.get("Makale/Yayın", 0), CENTER)
        dcell(ws5, ri, 12, tc.get("Bildiri", 0), CENTER)
        dcell(ws5, ri, 13, tc.get("Patent", 0), CENTER)
        dcell(ws5, ri, 14, oth, CENTER)

    col_ws5 = [14, 34, 16, 22, 24, 22, 14, 13, 13, 12, 10, 10, 10, 10]
    for i, w in enumerate(col_ws5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.row_dimensions[2].height = 32

    # ══════════════════════════════════════════════
    # SAYFA 6 — ÇIKTISI OLMAYAN PROJELER
    # ══════════════════════════════════════════════
    ws6 = wb.create_sheet("Ciktisi Olmayan Projeler")
    title_row(ws6, 1, f"Ciktisi Olmayan Projeler — {year_label}", 8)
    ws6.cell(row=2, column=1, value=(
        f"Toplam {len(data['no_output_projects'])} proje henuz cikti eklenmemis  |  "
        f"{len(data['no_output_declared_projects'])} proje 'Cikti Yok' bildirimi yapmis"
    )).font = Font(italic=True, color="718096")

    cols6 = ["Durum", "Proje Kodu", "Proje Basligi", "Proje Turu",
             "Yurutucu", "Fakulte", "Bolum", "Butce (TL)", "Bekleyen Cikti"]
    hrow(ws6, 3, cols6, color="C53030")
    ws6.freeze_panes = "A4"

    RED_FILL    = PatternFill("solid", fgColor="FFF5F5")
    PURPLE_FILL = PatternFill("solid", fgColor="FAF5FF")
    ri = 4

    for e in data["no_output_projects"]:
        durum = f"Beklemede ({e['pending']})" if e["pending"] else "Hic Eklenmemis"
        for col in range(1, 10):
            ws6.cell(row=ri, column=col).fill = RED_FILL
        dcell(ws6, ri, 1, durum, CENTER)
        dcell(ws6, ri, 2, e["code"], CENTER)
        dcell(ws6, ri, 3, e["title"])
        dcell(ws6, ri, 4, e["type"])
        dcell(ws6, ri, 5, e["pi"])
        dcell(ws6, ri, 6, e["faculty"])
        dcell(ws6, ri, 7, e["department"])
        dcell(ws6, ri, 8, e["budget"], CENTER)
        pend_c = dcell(ws6, ri, 9, e["pending"], CENTER)
        if e["pending"] > 0:
            pend_c.font = Font(bold=True, color="744210")
        ri += 1

    if data["no_output_declared_projects"]:
        ri += 1
        ws6.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=9)
        lbl = ws6.cell(row=ri, column=1, value="--- Cikti Yok Bildirimi Yapilan Projeler ---")
        lbl.font = Font(bold=True, color="553C9A")
        lbl.alignment = CENTER
        lbl.fill = PatternFill("solid", fgColor="EDE9FE")
        ri += 1
        hrow(ws6, ri, cols6, color="553C9A")
        ri += 1
        for e in data["no_output_declared_projects"]:
            for col in range(1, 10):
                ws6.cell(row=ri, column=col).fill = PURPLE_FILL
            dcell(ws6, ri, 1, "Cikti Yok Bildirimi", CENTER)
            dcell(ws6, ri, 2, e["code"], CENTER)
            dcell(ws6, ri, 3, e["title"])
            dcell(ws6, ri, 4, e["type"])
            dcell(ws6, ri, 5, e["pi"])
            dcell(ws6, ri, 6, e["faculty"])
            dcell(ws6, ri, 7, e["department"])
            dcell(ws6, ri, 8, e["budget"], CENTER)
            dcell(ws6, ri, 9, 0, CENTER)
            ri += 1

    col_ws6 = [20, 14, 36, 16, 24, 24, 22, 14, 14]
    for i, w in enumerate(col_ws6, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w
    ws6.row_dimensions[3].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
